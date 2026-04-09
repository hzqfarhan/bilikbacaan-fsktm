import openpyxl
import json

file_path = 'public/jadual/Jadual Semester 1 (2526).xlsm'
wb = openpyxl.load_workbook(file_path, data_only=True)

people = ['FARHAN', 'DANIAL', 'QAID', 'HANNAN', 'ZAAMAH']
time_slots = ['08:00-09:00', '09:00-10:00', '10:00-11:00', '11:00-12:00', '12:00-1:00', '1:00-2:00', '2:00-3:00', '3:00-4:00', '4:00-5:00', '5:00-6:00']
days = ['MON', 'TUE', 'WED', 'THU', 'FRI']

# Row 6=MON, 7=TUE, 8=WED, 9=THU, 10=FRI
# Col 2=08:00-09:00, 3=09:00-10:00, ... 11=5:00-6:00

timetable = {}
for person in people:
    if person not in wb.sheetnames:
        print(f"WARNING: Sheet '{person}' not found")
        continue
    sheet = wb[person]
    person_data = {}
    for r_idx, day in enumerate(days, start=6):
        day_data = {}
        for c_idx, slot in enumerate(time_slots, start=2):
            val = sheet.cell(row=r_idx, column=c_idx).value
            if val is not None and str(val).strip() != "":
                day_data[slot] = "BUSY"
            else:
                day_data[slot] = "FREE"
        person_data[day] = day_data
    timetable[person] = person_data

output = {
    "people": people,
    "days": days,
    "timeSlots": time_slots,
    "timetable": timetable
}

output_path = 'public/jadual/timetable.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print(f"Successfully generated {output_path}")
print(json.dumps(output, indent=2))
